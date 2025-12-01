# check_email.py
import csv
import io
import json
import socket
import smtplib
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from openpyxl import Workbook
import dns.resolver
import multiprocessing
import re
import threading
import os
import logging
from globals import progress_status

progress_lock = threading.Lock()

log = logging.getLogger(__name__)
if not log.handlers:
    ch = logging.StreamHandler()
    ch.setFormatter(logging.Formatter("%(asctime)s %(levelname)s %(message)s"))
    log.addHandler(ch)
log.setLevel(logging.DEBUG)

EMAIL_RE = re.compile(r"[^@]+@[^@]+\.[^@]+")


def load_settings():
    defaults = {
        "threads": 20,
        "min_threads": 5,
        "max_threads": 50,
        "dns_timeout": 6,
        "smtp_timeout": 8,
        "from_address": "verify@yourdomain.com",
        "assume_mx_valid": False
    }
    try:
        with open("settings.json", "r", encoding="utf-8") as f:
            cfg = json.load(f)
            for k, v in defaults.items():
                cfg.setdefault(k, v)
            return cfg
    except FileNotFoundError:
        log.warning("settings.json not found — using defaults")
        return defaults
    except Exception:
        log.exception("Failed to load settings.json — using defaults")
        return defaults


def get_mx_hosts(domain, timeout=8):
    try:
        resolver = dns.resolver.Resolver()
        resolver.timeout = timeout
        resolver.lifetime = timeout
        records = resolver.resolve(domain, "MX")
        hosts = [r.exchange.to_text(omit_final_dot=True) if hasattr(r.exchange, "to_text") else str(r.exchange) for r in records]
        log.debug("MX for %s -> %s", domain, hosts)
        return hosts
    except Exception as ex:
        log.debug("No MX for %s: %s", domain, ex)
        return []


def smtp_check_host(mx_host, email, cfg):
    from_address = cfg.get("from_address", "verify@yourdomain.com")
    timeout = cfg.get("smtp_timeout", 8)
    host = mx_host.rstrip(".")
    last_err = None
    for attempt in (1, 2):
        try:
            server = smtplib.SMTP(timeout=timeout)
            server.connect(host)
            server.helo("yourdomain.com")
            server.mail(from_address)
            code, msg = server.rcpt(email)
            try:
                server.quit()
            except Exception:
                pass
            log.debug("SMTP %s rcpt %s -> %s %s", host, email, code, msg)
            try:
                code = int(code)
            except Exception:
                return False, f"risky:unknown_code:{code}"
            # Map codes
            if code in (250, 251):
                return True, f"Accepted ({code})"
            if code == 550:
                return False, f"Rejected ({code})"
            # 4xx = temporary / greylist -> risky hint
            if 400 <= code < 500:
                return False, f"risky:temp_error({code})"
            return False, f"risky:unknown_response({code})"
        except (smtplib.SMTPServerDisconnected, smtplib.SMTPConnectError, socket.timeout, socket.error) as ex:
            last_err = ex
            if attempt == 1:
                time.sleep(0.8)
                continue
            return False, f"err:{repr(ex)}"
        except Exception as ex:
            last_err = ex
            return False, f"err:{repr(ex)}"
    return False, f"err:{repr(last_err)}"


def check_catch_all(mx_hosts, domain, cfg):
    test_email = f"nonexist_{int(time.time())}@{domain}"
    for host in mx_hosts:
        try:
            ok, _ = smtp_check_host(host, test_email, cfg)
            if ok:
                return True
        except Exception:
            continue
    return False


def verify_address(row, cfg):
    email_raw = (row.get("Email") or "").strip()
    name = row.get("Name") or ""
    result = {
        "Name": name,
        "Email": email_raw,
        "Status": "unknown",
        "Detail": "",
        "CheckedAt": datetime.utcnow().isoformat()
    }

    if not email_raw:
        result["Status"] = "invalid"
        result["Detail"] = "empty"
        return result

    email = email_raw.lower()
    if not EMAIL_RE.match(email):
        result["Status"] = "invalid"
        result["Detail"] = "bad_format"
        return result

    domain = email.split("@", 1)[1].lower()

    KNOWN_DELIVERABLES = {
        "gmail.com", "googlemail.com", "yahoo.com", "yahoo.co.in",
        "outlook.com", "hotmail.com", "live.com", "aol.com", "icloud.com",
        "msn.com", "protonmail.com", "me.com", "mac.com",
        "zoho.com", "office365.com", "gmx.com", "mail.com", "yandex.com"
    }

    if domain in KNOWN_DELIVERABLES:
        result["Status"] = "valid"
        result["Detail"] = "trusted_domain"
        return result

    mx_hosts = get_mx_hosts(domain, timeout=cfg.get("dns_timeout", 6))
    if not mx_hosts:
        result["Status"] = "invalid"
        result["Detail"] = "no_mx_records"
        return result

    accepted_any = False
    last_detail = ""
    risky_hint_found = False

    for host in mx_hosts:
        try:
            accepted, detail = smtp_check_host(host, email, cfg)
            last_detail = f"{host} - {detail}"
            if accepted:
                accepted_any = True
                break
            # remember risky hints but continue scanning other MXs
            if isinstance(detail, str) and detail.startswith("risky"):
                risky_hint_found = True
                # keep trying other MX records
                continue
        except Exception as ex:
            last_detail = f"{host} - err:{repr(ex)}"
            continue

    if accepted_any:
        try:
            is_catch = check_catch_all(mx_hosts, domain, cfg)
        except Exception:
            is_catch = False
        if is_catch:
            result["Status"] = "catchall"
            result["Detail"] = "catch_all_detected"
        else:
            result["Status"] = "valid"
            result["Detail"] = "rcpt_ok"
        return result

    # If none accepted, decide invalid or catch-all
    # Prefer marking invalid for explicit rejections or clear "no such user" responses
    # Treat risky hints as invalid by default (more conservative)
    if last_detail:
        ld = last_detail.lower()
        if "rejected" in ld or "no_mx_records" in ld:
            result["Status"] = "invalid"
            result["Detail"] = last_detail
            return result
        if "nosuchuser" in ld or "5.1.1" in ld or "the email account that you tried to reach does not exist" in ld:
            result["Status"] = "invalid"
            result["Detail"] = last_detail
            return result
        if risky_hint_found or "risky" in ld:
            # conservative: mark invalid if servers returned only risky/temporary responses
            result["Status"] = "invalid"
            result["Detail"] = last_detail
            return result

    # If we reached here and MX exists but nothing accepted:
    if not cfg.get("assume_mx_valid", False):
        result["Status"] = "invalid"
        result["Detail"] = "mx_but_no_rcpt"
        return result

    # fallback: assume valid only if explicitly configured
    result["Status"] = "valid"
    result["Detail"] = "assume_mx_valid"
    return result


def main(csv_path, progress_id=None, orig_filename=None):
    cfg = load_settings()

    tcfg = str(cfg.get("threads", "20"))
    if tcfg.lower() == "auto":
        cores = multiprocessing.cpu_count()
        cfg["threads"] = max(cfg.get("min_threads", 5),
                             min(cfg.get("max_threads", 50), cores * 5))

    threads = int(cfg.get("threads", 20))
    log.debug("Running with %s threads", threads)

    try:
        with open(csv_path, newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            if not reader.fieldnames:
                log.debug("DEBUG: No headers found in CSV at %s", csv_path)
                return {"valid": 0, "invalid": 0, "catchall": 0, "unknown": 0, "total": 0}, io.BytesIO()
            rows = []
            for row in reader:
                email = (
                    row.get("Email") or row.get("email") or
                    row.get("Email Address") or row.get("email address") or ""
                ).strip()
                name = row.get("Name") or row.get("name") or ""
                if email:
                    rows.append({"Name": name, "Email": email})
    except Exception as ex:
        log.exception("Failed to open/read CSV %s", csv_path)
        return {"valid": 0, "invalid": 0, "catchall": 0, "unknown": 0, "total": 0}, io.BytesIO()

    log.debug("DEBUG loaded rows: %s", len(rows))

    total = len(rows)
    if total == 0:
        log.debug("DEBUG: No valid rows found in CSV.")
        return {"valid": 0, "invalid": 0, "catchall": 0, "unknown": 0, "total": 0}, io.BytesIO()

    results = []
    stats = {"valid": 0, "catchall": 0, "invalid": 0, "unknown": 0}
    start_time = time.time()

    with ThreadPoolExecutor(max_workers=threads) as executor:
        futures = {executor.submit(verify_address, row, cfg): row for row in rows}
        completed_count = 0
        for future in as_completed(futures):
            completed_count += 1
            try:
                res = future.result()
            except Exception as e:
                log.exception("verify_address exception")
                res = {"Name": "", "Email": "", "Status": "invalid", "Detail": str(e), "CheckedAt": datetime.utcnow().isoformat()}
            results.append(res)

            st = res.get("Status", "invalid")
            if st == "valid":
                stats["valid"] += 1
            elif st == "catchall":
                stats["catchall"] += 1
            elif st == "unknown":
                stats["unknown"] += 1
            else:
                stats["invalid"] += 1

            if progress_id and (completed_count % 5 == 0 or completed_count == total):
                with progress_lock:
                    progress_status[progress_id] = {
                        "percent": int(completed_count / total * 100),
                        "verified": completed_count,
                        "queue": total - completed_count,
                        "state": "running",
                        "eta_seconds": max(1, int((time.time() - start_time) / max(1, completed_count) * (total - completed_count)))
                    }

    duration = round(time.time() - start_time, 2)
    log.debug("Processed %s emails in %ss", total, duration)

    wb = Workbook()
    categories = {
        "Valid": [r for r in results if r["Status"] == "valid"],
        "Risky": [r for r in results if r["Status"] == "catchall"],
        "Bad": [r for r in results if r["Status"] == "invalid"]
    }

    for sheet_name, data in categories.items():
        ws = wb.create_sheet(sheet_name)
        ws.append(["Name", "Email", "Status", "Detail", "CheckedAt"])
        for r in data:
            ws.append([r.get("Name", ""), r.get("Email", ""), r.get("Status", ""), r.get("Detail", ""), r.get("CheckedAt", "")])

    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    stats["total"] = total
    return stats, output
